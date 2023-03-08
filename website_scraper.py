import openpyxl
import re
from selenium import webdriver

# Open the notepad file and read the URLs
with open("websites.txt") as f:
    urls = f.read().splitlines()

# Initialize the Excel workbook and worksheet
workbook = openpyxl.Workbook()
worksheet = workbook.active

# Set the header row for the worksheet
header_row = ["Website", "Email", "Phone"]
worksheet.append(header_row)

# Initialize the WebDriver
driver = webdriver.Chrome()

# Loop through each URL
for i, url in enumerate(urls):
    # Check if the URL starts with "https://"
    if not url.startswith("https://"):
        url = "https://" + url

    # Navigate to the website
    driver.get(url)
    
    # Print the current website being visited
    print(f"Data Scrapper(1.0.0) {i+1}: {url}")

    # Get the HTML content of the page
    html = driver.page_source

    # Find the email address using a regular expression
    email_pattern = r'\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2,}\b'
    email_match = re.search(email_pattern, html)
    email = email_match.group(0) if email_match else ""

    # Find the phone number using a regular expression
    phone_pattern = r'\b\d{3}[-.]?\d{3}[-.]?\d{4}\b'
    phone_match = re.search(phone_pattern, html)
    phone = phone_match.group(0) if phone_match else ""

    # Check if the email and phone number are not empty
    if email != "" and phone != "":
        # Check if the row already exists in the worksheet
        row_exists = False
        for row in worksheet.iter_rows(min_row=2):
            if row[0].value == url and row[1].value == email and row[2].value == phone:
                row_exists = True
                break

        # Add a new row to the worksheet if the row does not already exist
        if not row_exists:
            row = [url, email, phone]
            worksheet.append(row)

# Save the results to an Excel file
workbook.save("results.xlsx")

# Print the summary
num_urls = len(urls)
num_results = worksheet.max_row - 1
print(f"Visited {num_urls} websites and found {num_results} results.")
